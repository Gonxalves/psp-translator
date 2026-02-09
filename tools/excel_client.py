"""
Excel Client

Handles read/write operations for Excel files stored in OneDrive/SharePoint.
Replacement for Google Sheets client - provides the same interface but works with local Excel files.
"""

import os
import time
from pathlib import Path
from typing import List, Optional
from dotenv import load_dotenv

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()


class ExcelFileLockError(Exception):
    """Raised when an Excel file is locked by another process."""
    pass


class ExcelFileNotFoundError(Exception):
    """Raised when an Excel file doesn't exist and can't be created."""
    pass


class ExcelClient:
    """
    Excel file client with file locking support.
    Replacement for GoogleSheetsClient for local/OneDrive Excel files.
    """

    def __init__(self):
        """Initialize the Excel client."""
        self.max_retries = int(os.getenv('EXCEL_MAX_RETRIES', '5'))
        self.retry_delay = float(os.getenv('EXCEL_RETRY_DELAY', '2'))

    def _is_file_locked(self, file_path: Path) -> bool:
        """
        Check if file is locked by Excel or another process.

        Excel creates a lock file named ~$filename.xlsx when open.
        On Linux servers, skip the lock file check since no Excel is running locally
        (lock files may be synced from OneDrive and are not relevant).
        """
        import sys

        # On Linux, only check for OS-level locks (not Excel lock files)
        if sys.platform != 'win32':
            if file_path.exists():
                try:
                    with open(file_path, 'r+b'):
                        pass
                    return False
                except PermissionError:
                    return True
            return False

        # Windows: check for Excel lock file
        lock_file = file_path.parent / f"~${file_path.name}"
        if lock_file.exists():
            return True

        # Try to open file exclusively to confirm it's not locked
        if file_path.exists():
            try:
                with open(file_path, 'r+b'):
                    pass
                return False
            except PermissionError:
                return True

        return False

    def _wait_for_unlock(self, file_path: Path) -> bool:
        """
        Wait for file to become available with retries.

        Returns:
            True if file became available, False if still locked after all retries
        """
        for attempt in range(self.max_retries):
            if not self._is_file_locked(file_path):
                return True

            wait_time = self.retry_delay * (attempt + 1)
            print(f"File locked, waiting {wait_time:.1f}s... (attempt {attempt + 1}/{self.max_retries})")
            time.sleep(wait_time)

        return False

    def _ensure_file_exists(self, file_path: Path, sheet_name: str, headers: List[str]):
        """
        Create Excel file with headers if it doesn't exist.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to create
            headers: List of header values for the first row
        """
        if file_path.exists():
            return

        # Ensure parent directory exists
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # Create new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(file_path)
        print(f"[OK] Created new Excel file: {file_path}")

    def read_sheet(self, file_path: Path, sheet_name: Optional[str] = None) -> List[List[str]]:
        """
        Read all data from an Excel sheet.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read (uses active sheet if None)

        Returns:
            List of rows, where each row is a list of cell values (as strings)

        Raises:
            ExcelFileLockError: If file is locked and doesn't become available
            ExcelFileNotFoundError: If file doesn't exist
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise ExcelFileNotFoundError(f"Excel file not found: {file_path}")

        # Wait for file to be available
        if self._is_file_locked(file_path):
            if not self._wait_for_unlock(file_path):
                raise ExcelFileLockError(
                    f"The file '{file_path.name}' is currently open in Excel. "
                    f"Please close it and try again."
                )

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return []
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Read all rows
            values = []
            for row in ws.iter_rows(values_only=True):
                # Convert None to empty string and all values to strings
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(cell.strip() for cell in row_values):
                    values.append(row_values)

            wb.close()
            return values

        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}")

    def append_row(self, file_path: Path, sheet_name: str, values: List[List[str]]) -> dict:
        """
        Append rows to an Excel sheet.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to append to
            values: List of rows to append, where each row is a list of cell values

        Returns:
            Dictionary with update details (rows_added, etc.)

        Raises:
            ExcelFileLockError: If file is locked
        """
        file_path = Path(file_path)

        # Wait for file to be available
        if self._is_file_locked(file_path):
            if not self._wait_for_unlock(file_path):
                raise ExcelFileLockError(
                    f"The file '{file_path.name}' is currently open in Excel. "
                    f"Please close it and try again."
                )

        try:
            if not file_path.exists():
                raise ExcelFileNotFoundError(f"Excel file not found: {file_path}")

            wb = load_workbook(file_path)

            if sheet_name not in wb.sheetnames:
                # Create the sheet if it doesn't exist
                wb.create_sheet(sheet_name)

            ws = wb[sheet_name]

            # Append each row
            rows_added = 0
            for row in values:
                ws.append(row)
                rows_added += 1

            wb.save(file_path)
            wb.close()

            return {
                'spreadsheetId': str(file_path),
                'updates': {
                    'updatedRows': rows_added
                }
            }

        except ExcelFileLockError:
            raise
        except Exception as e:
            raise Exception(f"Error appending to Excel file: {e}")

    def update_cell(self, file_path: Path, sheet_name: str, row: int, col: int, value: str) -> dict:
        """
        Update a specific cell in an Excel sheet.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            value: New value for the cell

        Returns:
            Dictionary with update details
        """
        file_path = Path(file_path)

        # Wait for file to be available
        if self._is_file_locked(file_path):
            if not self._wait_for_unlock(file_path):
                raise ExcelFileLockError(
                    f"The file '{file_path.name}' is currently open in Excel. "
                    f"Please close it and try again."
                )

        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            ws.cell(row=row, column=col, value=value)
            wb.save(file_path)
            wb.close()

            return {
                'spreadsheetId': str(file_path),
                'updatedCells': 1
            }

        except Exception as e:
            raise Exception(f"Error updating cell: {e}")

    def batch_update(self, file_path: Path, sheet_name: str, updates: List[dict]) -> dict:
        """
        Perform multiple update operations in a single file open/save.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            updates: List of update dictionaries with 'row', 'col', and 'value' keys

        Returns:
            Dictionary with update details
        """
        file_path = Path(file_path)

        # Wait for file to be available
        if self._is_file_locked(file_path):
            if not self._wait_for_unlock(file_path):
                raise ExcelFileLockError(
                    f"The file '{file_path.name}' is currently open in Excel. "
                    f"Please close it and try again."
                )

        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            cells_updated = 0
            for update in updates:
                row = update['row']
                col = update['col']
                value = update['value']
                ws.cell(row=row, column=col, value=value)
                cells_updated += 1

            wb.save(file_path)
            wb.close()

            return {
                'spreadsheetId': str(file_path),
                'updatedCells': cells_updated
            }

        except Exception as e:
            raise Exception(f"Error in batch update: {e}")


# Singleton instance for reuse across the application
_client_instance = None


def get_client() -> ExcelClient:
    """
    Get or create a singleton instance of the Excel client.

    Returns:
        ExcelClient instance
    """
    global _client_instance
    if _client_instance is None:
        _client_instance = ExcelClient()
    return _client_instance


# File path helpers
def get_glossary_path() -> Path:
    """Get the glossary Excel file path from environment."""
    path = os.getenv('EXCEL_GLOSSARY_PATH', '')
    if not path:
        raise ValueError(
            "EXCEL_GLOSSARY_PATH not set in .env file. "
            "Please add the path to your glossary Excel file."
        )
    return Path(path)


def get_action_log_path() -> Path:
    """Get the action log Excel file path from environment."""
    path = os.getenv('EXCEL_ACTION_LOG_PATH', '')
    if not path:
        raise ValueError(
            "EXCEL_ACTION_LOG_PATH not set in .env file. "
            "Please add the path to your action log Excel file."
        )
    return Path(path)


def ensure_glossary_exists():
    """Create glossary file with headers if it doesn't exist."""
    client = get_client()
    path = get_glossary_path()
    client._ensure_file_exists(
        path,
        sheet_name="Glossary",
        headers=["French Term", "English Term", "Notes"]
    )


def ensure_action_log_exists():
    """Create action log file with headers if it doesn't exist."""
    client = get_client()
    path = get_action_log_path()
    client._ensure_file_exists(
        path,
        sheet_name="Action Log",
        headers=["Timestamp", "French Term", "English Term", "Source", "Added to Glossary"]
    )


if __name__ == "__main__":
    # Test the client
    print("Testing Excel Client...")
    try:
        client = get_client()
        print("[OK] Excel client initialized!")

        # Check if paths are configured
        try:
            glossary_path = get_glossary_path()
            print(f"[OK] Glossary path: {glossary_path}")

            if glossary_path.exists():
                print("[OK] Glossary file exists")
                values = client.read_sheet(glossary_path, "Glossary")
                print(f"[OK] Read {len(values)} rows from glossary")
            else:
                print("[INFO] Glossary file does not exist yet")

        except ValueError as e:
            print(f"[WARNING] {e}")

        try:
            action_log_path = get_action_log_path()
            print(f"[OK] Action log path: {action_log_path}")
        except ValueError as e:
            print(f"[WARNING] {e}")

    except Exception as e:
        print(f"[ERROR] {e}")
